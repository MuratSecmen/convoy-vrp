"""
src/export.py
=============
Export Pareto frontier and solution details to Excel and JSON.
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from src.model import Solution


# ── Colour palette ────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1F3864")  # dark navy
ROW_FILL  = PatternFill("solid", start_color="DCE6F1")  # light blue
ALT_FILL  = PatternFill("solid", start_color="FFFFFF")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=9)
KNEE_FILL = PatternFill("solid", start_color="FFC000")  # amber knee


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = HDR_FONT
    c.fill  = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border = _border()
    return c


def _cell(ws, row, col, val, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = BODY_FONT
    c.fill   = fill or ALT_FILL
    c.border = _border()
    c.alignment = Alignment(horizontal="center")
    if fmt:
        c.number_format = fmt
    return c


def _knee_index(solutions: List[Solution]) -> int:
    """Maximum curvature (angle) on Pareto frontier."""
    if len(solutions) < 3:
        return 0
    w1 = [s.W1 for s in solutions]
    w2 = [s.W2 for s in solutions]
    # Normalise
    r1 = max(w1) - min(w1) or 1
    r2 = max(w2) - min(w2) or 1
    angles = []
    for i in range(1, len(solutions)-1):
        a = ((w1[i]-w1[i-1])/r1, (w2[i]-w2[i-1])/r2)
        b = ((w1[i+1]-w1[i])/r1, (w2[i+1]-w2[i])/r2)
        dot   = a[0]*b[0] + a[1]*b[1]
        na    = (a[0]**2 + a[1]**2)**0.5 or 1
        nb    = (b[0]**2 + b[1]**2)**0.5 or 1
        angle = dot / (na * nb)
        angles.append(angle)
    return angles.index(min(angles)) + 1


def export_results(solutions: List[Solution],
                   out_path: str = "results/pareto.xlsx"):
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════
    # Sheet 1 — Pareto Frontier
    # ════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Pareto Frontier"

    feasible = [s for s in solutions if s.feasible]
    knee_idx = _knee_index(feasible) if len(feasible) >= 3 else 0

    headers = ["Point", "Epsilon (W2 bound)", "W1 (max travel, min)",
               "W2 (max deviation)", "Knee?",
               "Status", "Solve Time (s)"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)

    for r, sol in enumerate(feasible, 2):
        fill = KNEE_FILL if (r-2) == knee_idx else (
            ROW_FILL if r % 2 == 0 else ALT_FILL)
        _cell(ws, r, 1, r-1, fill)
        _cell(ws, r, 2, round(sol.epsilon, 4), fill, "0.0000")
        _cell(ws, r, 3, sol.W1, fill, "0.00")
        _cell(ws, r, 4, sol.W2, fill, "0.0000")
        _cell(ws, r, 5, "★ KNEE" if (r-2) == knee_idx else "", fill)
        _cell(ws, r, 6, sol.status, fill)
        _cell(ws, r, 7, sol.solve_time, fill, "0.00")

    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # Sheet 2 — Vehicle Routes (knee point solution)
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Routes (Knee Point)")
    sol = feasible[knee_idx] if feasible else None

    if sol:
        ws2["A1"] = f"Instance: {sol.instance_name}"
        ws2["A1"].font = Font(bold=True, size=12, name="Arial")
        ws2["A2"] = (f"W1 = {sol.W1:.2f} min  |  "
                     f"W2 = {sol.W2:.4f}  |  Status: {sol.status}")
        ws2["A2"].font = Font(italic=True, size=10, name="Arial")

        route_hdrs = ["Vehicle", "Arc From", "Arc To", "Travel Time (min)",
                      "Actual L_k (min)", "Baseline L_bar (min)",
                      "Deviation (%)"]
        for col, h in enumerate(route_hdrs, 1):
            _hdr(ws2, 4, col, h)

        row = 5
        active_arcs = {(i,j,k): v for (i,j,k), v in sol.x.items()
                       if v == 1}
        for (i,j,k), _ in sorted(active_arcs.items()):
            fill = ROW_FILL if row % 2 == 0 else ALT_FILL
            lk   = sol.L.get(k, 0)
            _cell(ws2, row, 1, f"Vehicle {k}", fill)
            _cell(ws2, row, 2, i, fill)
            _cell(ws2, row, 3, j, fill)
            _cell(ws2, row, 4, "", fill)   # filled from travel_time
            _cell(ws2, row, 5, round(lk, 2), fill, "0.00")
            _cell(ws2, row, 6, "", fill)
            _cell(ws2, row, 7, f"={get_column_letter(5)}{row}-"
                               f"{get_column_letter(6)}{row}", fill, "0.00%")
            row += 1

        for col in range(1, len(route_hdrs)+1):
            ws2.column_dimensions[get_column_letter(col)].width = 22
        ws2.freeze_panes = "A5"

    # ════════════════════════════════════════════════════════════════
    # Sheet 3 — Service Times
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Service Times")
    if sol:
        svc_hdrs = ["Vehicle", "Node", "Service Time (hr)"]
        for col, h in enumerate(svc_hdrs, 1):
            _hdr(ws3, 1, col, h)
        row = 2
        for (k, i), v in sorted(sol.T.items()):
            if v > 1e-6:
                fill = ROW_FILL if row % 2 == 0 else ALT_FILL
                _cell(ws3, row, 1, f"Vehicle {k}", fill)
                _cell(ws3, row, 2, i, fill)
                _cell(ws3, row, 3, round(v, 4), fill, "0.0000")
                row += 1
        for col in range(1, 4):
            ws3.column_dimensions[get_column_letter(col)].width = 22
        ws3.freeze_panes = "A2"

    wb.save(out_path)
    print(f"  ✓ Excel results saved: {out_path}")

    # ── JSON dump ────────────────────────────────────────────────
    json_path = out_path.replace(".xlsx", ".json")
    data = []
    for s in feasible:
        data.append({
            "instance":   s.instance_name,
            "epsilon":    s.epsilon,
            "W1":         s.W1,
            "W2":         s.W2,
            "L":          s.L,
            "status":     s.status,
            "solve_time": s.solve_time,
        })
    with open(json_path, "w") as jf:
        json.dump(data, jf, indent=2)
    print(f"  ✓ JSON results saved:  {json_path}")
