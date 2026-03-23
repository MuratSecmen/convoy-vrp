"""
data/generate_stochastic_instances.py
======================================
Generates three stochastic benchmark Excel instances:
  small_stoch_n5_k3    : 5 nodes, 3 vehicles, 6 scenarios
  medium_stoch_n10_k5  : 10 nodes, 5 vehicles, 6 scenarios
  large_stoch_n15_k7   : 15 nodes, 7 vehicles, 6 scenarios

Sheets per workbook:
  Nodes | Vehicles | TravelTime | Demand | ServiceRate |
  Baseline | Regions | Scenarios | CongestionMultipliers
"""

import os
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

SUPPLY_CLASSES = ["I", "II", "III_W", "IV", "VIII", "IX"]

PLATFORMS = {
    0: {"name": "LMTV",          "capacity": 2.27,  "speed": 60},
    1: {"name": "MTV",           "capacity": 4.54,  "speed": 60},
    2: {"name": "HEMTT",         "capacity": 11.3,  "speed": 65},
    3: {"name": "HEMTT_Tanker",  "capacity": 9.46,  "speed": 60},
    4: {"name": "HEMTT_LHS",     "capacity": 11.3,  "speed": 55},
    5: {"name": "LMTV_Med",      "capacity": 2.27,  "speed": 60},
    6: {"name": "HEMTT_Wrecker", "capacity": 8.0,   "speed": 55},
}

VEH_CLASS = {
    "I":     [0, 1, 2],
    "II":    [0, 1, 2],
    "III_W": [3],
    "IV":    [4],
    "VIII":  [5],
    "IX":    [2, 6],
}

SERVICE_RATES = {
    (0,"I"):10,(1,"I"):15,(2,"I"):20,
    (0,"II"):8,(1,"II"):12,(2,"II"):18,
    (3,"III_W"):25,(4,"IV"):20,(5,"VIII"):5,
    (2,"IX"):10,(6,"IX"):12,
}

# ── Six calibrated scenarios (ISAF 2009-2013) ─────────────────────
SCENARIOS = [
    {"omega_id":1,"label":"Nominal",
     "probability":0.217,"blocked_arcs":"","delta":1.00,"kappa":0.0,
     "source":"Baseline operations; no disruption"},
    {"omega_id":2,"label":"Southern MSR closure (IED/VBIED)",
     "probability":0.196,"blocked_arcs":"see sheet",
     "delta":1.10,"kappa":0.2,
     "source":"RC-South 40% of all ISAF IED events (CFC 2013)"},
    {"omega_id":3,"label":"Eastern MSR closure (IED/tactical hold)",
     "probability":0.152,"blocked_arcs":"see sheet",
     "delta":1.00,"kappa":0.1,
     "source":"RC-East 33% of all ISAF IED events (CFC 2013)"},
    {"omega_id":4,"label":"Demand surge",
     "probability":0.152,"blocked_arcs":"",
     "delta":1.35,"kappa":0.0,
     "source":"Emergency resupply; SIGAR logistics reports"},
    {"omega_id":5,"label":"ROE restriction (dual-command)",
     "probability":0.130,"blocked_arcs":"see sheet",
     "delta":1.00,"kappa":1.0,
     "source":"ISAF/CJTF re-authorisation events; AJP-4.4"},
    {"omega_id":6,"label":"Compound disruption (IED + surge)",
     "probability":0.152,"blocked_arcs":"see sheet",
     "delta":1.20,"kappa":0.5,
     "source":"Salala-type compound closure pattern"},
]

# ── Styling ───────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", start_color="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FILL = PatternFill("solid", start_color="DCE6F1")
ALT_FILL = PatternFill("solid", start_color="FFFFFF")
SCN_FILL = PatternFill("solid", start_color="FFF2CC")  # amber for scenarios
BODY_FONT = Font(name="Arial", size=9)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_df(ws, df, title=None):
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=11,
                              color="1F3864", name="Arial")
        start = 3
    else:
        start = 1
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=start, column=ci, value=col)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.border=_border()
        c.alignment=Alignment(horizontal="center", wrap_text=True)
    for ri, row in enumerate(df.itertuples(index=False), start+1):
        fill = ROW_FILL if ri%2==0 else ALT_FILL
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font=BODY_FONT; c.fill=fill; c.border=_border()
            c.alignment=Alignment(horizontal="center")
    for ci in range(1, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    ws.freeze_panes = f"A{start+1}"


def generate(n_del: int, n_veh: int, seed: int, name: str) -> Workbook:
    rng  = np.random.default_rng(seed)
    wb   = Workbook()
    dep  = 0
    nodes = list(range(n_del+1))
    vehs  = list(range(n_veh))
    plats = list(PLATFORMS.keys())

    coords = {0: (50.0, 50.0)}
    for i in range(1, n_del+1):
        coords[i] = (rng.uniform(10,90), rng.uniform(10,90))

    n_reg = max(2, n_del//4)
    region = {0:0}
    for i in range(1, n_del+1):
        region[i] = int(rng.integers(1, n_reg+1))

    # ── Nodes ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Nodes"
    types = ["depot"]+["COP" if i%3==1 else "VSP" if i%3==2
                        else "ANP" for i in range(1,n_del+1)]
    df_n = pd.DataFrame({
        "node_id": nodes,
        "type":    types,
        "x_km":    [round(coords[i][0],2) for i in nodes],
        "y_km":    [round(coords[i][1],2) for i in nodes],
        "ao_region": [region[i] for i in nodes],
    })
    _write_df(ws, df_n, f"Nodes — {name}")

    # ── Vehicles ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Vehicles")
    df_v = pd.DataFrame({
        "vehicle_id":      vehs,
        "platform":        [PLATFORMS[plats[k%len(plats)]]["name"]
                            for k in vehs],
        "capacity_tonnes": [PLATFORMS[plats[k%len(plats)]]["capacity"]
                            for k in vehs],
        "speed_paved_kmh": [PLATFORMS[plats[k%len(plats)]]["speed"]
                            for k in vehs],
        "observable":      [1 if k%2==0 else 0 for k in vehs],
    })
    _write_df(ws2, df_v, f"Vehicles — {name}")

    # ── TravelTime ────────────────────────────────────────────────
    ws3 = wb.create_sheet("TravelTime")
    M   = 9999.0
    rows_tt = []
    for k in vehs:
        sp = PLATFORMS[plats[k%len(plats)]]["speed"]
        for i in nodes:
            for j in nodes:
                if i==j: continue
                dx,dy = coords[i][0]-coords[j][0], coords[i][1]-coords[j][1]
                d = np.hypot(dx,dy)
                blocked = (i!=dep and j!=dep and
                            region[i]!=region[j])
                tt = M if blocked else round(d/sp*60+rng.uniform(3,12),1)
                rows_tt.append({
                    "from":i,"to":j,"vehicle_id":k,
                    "time_min":tt,
                    "note":"BLOCKED(inter-AO)" if tt>=M else "normal",
                })
    _write_df(ws3, pd.DataFrame(rows_tt), f"Travel Time (min) — {name}")

    # ── Demand ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Demand")
    rows_d = []
    for i in range(1, n_del+1):
        for d in SUPPLY_CLASSES:
            qty = round(rng.uniform(5,25),1) if rng.random()<0.5 else 0.0
            rows_d.append({"node_id":i,"class":d,
                           "quantity":qty,"unit":"service_units"})
    _write_df(ws4, pd.DataFrame(rows_d), f"Demand — {name}")

    # ── ServiceRate ───────────────────────────────────────────────
    ws5 = wb.create_sheet("ServiceRate")
    rows_sr = []
    for k in vehs:
        pid = plats[k%len(plats)]
        for d in SUPPLY_CLASSES:
            rows_sr.append({
                "vehicle_id":k,
                "platform":PLATFORMS[pid]["name"],
                "class":d,
                "rate":SERVICE_RATES.get((pid,d),0.0),
                "unit":"service_units/hour",
            })
    _write_df(ws5, pd.DataFrame(rows_sr), f"Service Rate — {name}")

    # ── Baseline ──────────────────────────────────────────────────
    ws6 = wb.create_sheet("Baseline")
    rows_bl = []
    for k in vehs:
        sp  = PLATFORMS[plats[k%len(plats)]]["speed"]
        avg = rng.uniform(30,70)
        bl  = round(avg/sp*60+rng.uniform(20,40),1)
        rows_bl.append({
            "vehicle_id":k,
            "platform":PLATFORMS[plats[k%len(plats)]]["name"],
            "baseline_min":bl,
            "note":f"CJ4 morning briefing plan — vehicle {k}",
        })
    _write_df(ws6, pd.DataFrame(rows_bl), f"Baseline Plan — {name}")

    # ── Regions ───────────────────────────────────────────────────
    ws7 = wb.create_sheet("Regions")
    df_r = pd.DataFrame({
        "node_id":  nodes,
        "ao_region":[region[i] for i in nodes],
        "label":    [f"AO-{region[i]}" if i!=dep else "DEPOT"
                     for i in nodes],
    })
    _write_df(ws7, df_r, f"Areas of Operation — {name}")

    # ── Scenarios ─────────────────────────────────────────────────
    ws8 = wb.create_sheet("Scenarios")
    df_s = pd.DataFrame(SCENARIOS)
    # colour scenario rows amber
    _write_df(ws8, df_s, f"Disruption Scenarios — {name}")
    # re-colour body amber for scenarios sheet
    for r in range(4, 4+len(SCENARIOS)):
        for c in range(1, len(df_s.columns)+1):
            ws8.cell(row=r, column=c).fill = SCN_FILL

    # ── CongestionMultipliers ─────────────────────────────────────
    ws9 = wb.create_sheet("CongestionMultipliers")
    rows_mu = []
    for o in SCENARIOS:
        oid = o["omega_id"]
        for k in vehs:
            for i in nodes:
                for j in nodes:
                    if i==j: continue
                    # mu=1 nominal, mu>1 congested arcs in disruption
                    if oid == 1:
                        mu = 1.0
                    elif oid in [2,3,6]:
                        mu = round(rng.uniform(1.1, 1.5), 2)
                    else:
                        mu = round(rng.uniform(1.0, 1.2), 2)
                    rows_mu.append({
                        "omega_id":oid,
                        "from":i,"to":j,"vehicle_id":k,
                        "mu":mu,
                    })
    _write_df(ws9, pd.DataFrame(rows_mu),
              f"Arc Congestion Multipliers (mu) — {name}")

    # ── BlockedArcs ───────────────────────────────────────────────
    ws10 = wb.create_sheet("BlockedArcs")
    rows_ba = []
    # For scenarios 2,3,5,6 block some arcs based on region geography
    # Southern = high region numbers, Eastern = mid region numbers
    south_nodes = [i for i in range(1,n_del+1) if region[i] >= n_reg]
    east_nodes  = [i for i in range(1,n_del+1)
                   if 1 < region[i] < n_reg]

    def add_blocks(oid, nodeset):
        for i in nodeset:
            for j in nodeset:
                if i!=j:
                    rows_ba.append({
                        "omega_id":oid,"from":i,"to":j,
                        "note":"blocked arc"
                    })

    add_blocks(2, south_nodes)
    add_blocks(3, east_nodes)
    add_blocks(5, south_nodes[:max(1,len(south_nodes)//2)])
    add_blocks(6, south_nodes+east_nodes)

    if rows_ba:
        _write_df(ws10, pd.DataFrame(rows_ba),
                  f"Blocked Arcs per Scenario — {name}")
    else:
        ws10["A1"] = "No blocked arcs defined (all AOs connected)"

    return wb


def main():
    configs = [
        (5,  3, 42,  "small_stoch_n5_k3"),
        (10, 5, 7,   "medium_stoch_n10_k5"),
        (15, 7, 99,  "large_stoch_n15_k7"),
    ]
    os.makedirs("data/instances", exist_ok=True)
    for n,k,seed,name in configs:
        wb   = generate(n,k,seed,name)
        path = f"data/instances/{name}.xlsx"
        wb.save(path)
        print(f"  ✓ {path}")


if __name__ == "__main__":
    main()
