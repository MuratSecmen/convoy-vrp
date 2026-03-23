"""
data/generate_instances.py
==========================
Generates three benchmark Excel instances for BCMC-VRPHD:
  - small_n5_k3  :  5 delivery nodes, 3 vehicles
  - medium_n10_k5: 10 delivery nodes, 5 vehicles
  - large_n15_k7 : 15 delivery nodes, 7 vehicles

Each workbook contains sheets:
  Nodes | Vehicles | TravelTime | Demand |
  ServiceRate | Baseline | Regions
"""

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPPLY_CLASSES = ["I", "II", "III_W", "IV", "VIII", "IX"]

# Platform data (ISAF ground vehicles)
PLATFORMS = {
    0: {"name": "LMTV",         "capacity": 2.27,  "speed_paved": 60},
    1: {"name": "MTV",          "capacity": 4.54,  "speed_paved": 60},
    2: {"name": "HEMTT",        "capacity": 11.3,  "speed_paved": 65},
    3: {"name": "HEMTT_Tanker", "capacity": 9.46,  "speed_paved": 60},
    4: {"name": "HEMTT_LHS",    "capacity": 11.3,  "speed_paved": 55},
    5: {"name": "LMTV_Med",     "capacity": 2.27,  "speed_paved": 60},
    6: {"name": "HEMTT_Wrecker","capacity": 8.0,   "speed_paved": 55},
}

# Vehicle-class assignments (which platform can carry which class)
VEH_CLASS_COMPAT = {
    "I":     [0, 1, 2],        # LMTV, MTV, HEMTT
    "II":    [0, 1, 2],
    "III_W": [3],               # HEMTT Tanker only
    "IV":    [4],               # HEMTT LHS only
    "VIII":  [5],               # LMTV Medical only
    "IX":    [2, 6],            # HEMTT, HEMTT Wrecker
}

# Service rates (service units / hour) per vehicle per class
SERVICE_RATES = {
    (0, "I"):     10.0,
    (1, "I"):     15.0,
    (2, "I"):     20.0,
    (0, "II"):    8.0,
    (1, "II"):    12.0,
    (2, "II"):    18.0,
    (3, "III_W"): 25.0,
    (4, "IV"):    20.0,
    (5, "VIII"):  5.0,
    (2, "IX"):    10.0,
    (6, "IX"):    12.0,
}

HDR_FILL = PatternFill("solid", start_color="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FILL = PatternFill("solid", start_color="DCE6F1")
ALT_FILL = PatternFill("solid", start_color="FFFFFF")
BODY_FONT = Font(name="Arial", size=9)


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_df(ws, df, title=None):
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=11,
                              color="1F3864", name="Arial")
        start_row = 3
    else:
        start_row = 1

    for col_idx, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=col_idx, value=col_name)
        c.font   = HDR_FONT
        c.fill   = HDR_FILL
        c.border = _border()
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx, row in enumerate(df.itertuples(index=False),
                                 start_row + 1):
        fill = ROW_FILL if r_idx % 2 == 0 else ALT_FILL
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font   = BODY_FONT
            c.fill   = fill
            c.border = _border()
            c.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = (f"A{start_row + 1}" if title
                       else "A2")


def generate_instance(n_delivery: int,
                      n_vehicles: int,
                      seed: int,
                      name: str) -> Workbook:
    rng = np.random.default_rng(seed)
    wb  = Workbook()

    depot_id = 0
    node_ids = list(range(0, n_delivery + 1))
    veh_ids  = list(range(n_vehicles))

    # ── Coordinates (Euclidean, km) ──────────────────────────────
    # Depot at centre, delivery nodes around it
    coords = {depot_id: (50.0, 50.0)}
    for i in range(1, n_delivery + 1):
        coords[i] = (rng.uniform(10, 90), rng.uniform(10, 90))

    # ── Regions ──────────────────────────────────────────────────
    n_regions = max(2, n_delivery // 4)
    region_assign = {0: 0}
    for i in range(1, n_delivery + 1):
        region_assign[i] = int(rng.integers(1, n_regions + 1))

    # ─────────────────────────────────────────────────────────────
    # Sheet: Nodes
    # ─────────────────────────────────────────────────────────────
    ws_nodes = wb.active
    ws_nodes.title = "Nodes"
    node_types = (["depot"] +
                  ["COP" if i % 3 == 1 else
                   "VSP" if i % 3 == 2 else
                   "ANP" for i in range(1, n_delivery + 1)])
    df_nodes = pd.DataFrame({
        "node_id": node_ids,
        "type":    node_types,
        "x_km":    [round(coords[i][0], 2) for i in node_ids],
        "y_km":    [round(coords[i][1], 2) for i in node_ids],
        "description": (
            ["Staging Area / FOB"] +
            [f"Delivery Node {i}" for i in range(1, n_delivery + 1)]
        ),
    })
    _write_df(ws_nodes, df_nodes, f"Nodes — {name}")

    # ─────────────────────────────────────────────────────────────
    # Sheet: Vehicles
    # ─────────────────────────────────────────────────────────────
    ws_veh = wb.create_sheet("Vehicles")
    plat_cycle = list(PLATFORMS.keys())
    df_veh = pd.DataFrame({
        "vehicle_id":      veh_ids,
        "platform":        [PLATFORMS[plat_cycle[k % len(plat_cycle)]]
                            ["name"] for k in veh_ids],
        "capacity_tonnes": [PLATFORMS[plat_cycle[k % len(plat_cycle)]]
                            ["capacity"] for k in veh_ids],
        "speed_paved_kmh": [PLATFORMS[plat_cycle[k % len(plat_cycle)]]
                            ["speed_paved"] for k in veh_ids],
        "observable":      [1 if k % 2 == 0 else 0 for k in veh_ids],
    })
    _write_df(ws_veh, df_veh, f"Vehicles — {name}")

    # ─────────────────────────────────────────────────────────────
    # Sheet: TravelTime
    # ─────────────────────────────────────────────────────────────
    ws_tt = wb.create_sheet("TravelTime")
    M_big = 9999.0  # inter-region blocking cost
    rows_tt = []
    for k in veh_ids:
        plat = PLATFORMS[plat_cycle[k % len(plat_cycle)]]
        spd  = plat["speed_paved"]
        for i in node_ids:
            for j in node_ids:
                if i == j:
                    continue
                dx  = coords[i][0] - coords[j][0]
                dy  = coords[i][1] - coords[j][1]
                dist = np.hypot(dx, dy)
                # Block inter-region arcs (not depot)
                if (i != depot_id and j != depot_id and
                        region_assign[i] != region_assign[j]):
                    tt = M_big
                else:
                    tt = round((dist / spd) * 60 +         # travel
                               rng.uniform(3, 12), 1)       # checkpoint
                rows_tt.append({
                    "from":       i,
                    "to":         j,
                    "vehicle_id": k,
                    "time_min":   tt,
                    "note":       ("BLOCKED" if tt >= M_big
                                   else "normal"),
                })
    df_tt = pd.DataFrame(rows_tt)
    _write_df(ws_tt, df_tt, f"Travel Time (minutes) — {name}")

    # ─────────────────────────────────────────────────────────────
    # Sheet: Demand
    # ─────────────────────────────────────────────────────────────
    ws_dem = wb.create_sheet("Demand")
    rows_dem = []
    for i in range(1, n_delivery + 1):
        for d in SUPPLY_CLASSES:
            # Each node demands 1-3 classes, others zero
            qty = round(rng.uniform(5, 25), 1) if rng.random() < 0.5 else 0.0
            rows_dem.append({
                "node_id":  i,
                "class":    d,
                "quantity": qty,
                "unit":     "service_units",
                "note":     f"Class {d} demand at node {i}",
            })
    df_dem = pd.DataFrame(rows_dem)
    _write_df(ws_dem, df_dem, f"Demand — {name}")

    # ─────────────────────────────────────────────────────────────
    # Sheet: ServiceRate
    # ─────────────────────────────────────────────────────────────
    ws_sr = wb.create_sheet("ServiceRate")
    rows_sr = []
    for k in veh_ids:
        plat_id = plat_cycle[k % len(plat_cycle)]
        for d in SUPPLY_CLASSES:
            r = SERVICE_RATES.get((plat_id, d), 0.0)
            rows_sr.append({
                "vehicle_id":       k,
                "platform":         PLATFORMS[plat_id]["name"],
                "class":            d,
                "rate":             r,
                "unit":             "service_units/hour",
            })
    df_sr = pd.DataFrame(rows_sr)
    _write_df(ws_sr, df_sr, f"Service Rate — {name}")

    # ─────────────────────────────────────────────────────────────
    # Sheet: Baseline
    # ─────────────────────────────────────────────────────────────
    ws_base = wb.create_sheet("Baseline")
    rows_base = []
    for k in veh_ids:
        plat = PLATFORMS[plat_cycle[k % len(plat_cycle)]]
        spd  = plat["speed_paved"]
        # Baseline = expected route length / speed + service buffer
        avg_dist = rng.uniform(30, 70)
        bl = round((avg_dist / spd) * 60 + rng.uniform(20, 40), 1)
        rows_base.append({
            "vehicle_id":    k,
            "platform":      plat["name"],
            "baseline_min":  bl,
            "note": ("CJ4 morning briefing approved plan "
                     f"for vehicle {k}"),
        })
    df_base = pd.DataFrame(rows_base)
    _write_df(ws_base, df_base, f"Baseline Plan — {name}")

    # ─────────────────────────────────────────────────────────────
    # Sheet: Regions
    # ─────────────────────────────────────────────────────────────
    ws_reg = wb.create_sheet("Regions")
    rows_reg = []
    for i in node_ids:
        rows_reg.append({
            "node_id":   i,
            "ao_region": region_assign[i],
            "label":     (f"AO-{region_assign[i]}"
                          if i != depot_id else "DEPOT"),
        })
    df_reg = pd.DataFrame(rows_reg)
    _write_df(ws_reg, df_reg, f"Areas of Operation — {name}")

    return wb


def main():
    configs = [
        (5,  3, 42,  "small_n5_k3"),
        (10, 5, 7,   "medium_n10_k5"),
        (15, 7, 99,  "large_n15_k7"),
    ]
    for n, k, seed, name in configs:
        wb   = generate_instance(n, k, seed, name)
        os.makedirs("data/instances", exist_ok=True)
        path = f"data/instances/{name}.xlsx"
        wb.save(path)
        print(f"  ✓ Generated: {path}")


if __name__ == "__main__":
    main()
