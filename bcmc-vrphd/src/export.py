"""
BCMC-VRPHD Result Exporter
============================
Exports Pareto frontier to Excel (with knee-point highlighting) and JSON.
"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, List


def export_results(
    frontier: List[Dict],
    instance_name: str,
    output_dir: str = "results",
):
    """Export Pareto frontier to Excel and JSON."""
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    # ── Pareto table ─────────────────────────────────────────
    rows = []
    for idx, sol in enumerate(frontier):
        row = {
            "point": idx + 1,
            "W1": round(sol["W1"], 2),
            "W2": round(sol["W2"], 6),
            "status": sol["status"],
            "knee": sol.get("knee", False),
        }
        # Add per-vehicle L_k
        for k, lk in sol.get("L", {}).items():
            row[f"L_{k}"] = round(lk, 2) if lk is not None else None
        rows.append(row)

    df = pd.DataFrame(rows)

    # ── Route details ────────────────────────────────────────
    route_rows = []
    for idx, sol in enumerate(frontier):
        for k, route in sol.get("routes", {}).items():
            route_rows.append({
                "point": idx + 1,
                "vehicle": k,
                "route": " → ".join(str(n) for n in route),
                "L_k": round(sol["L"].get(k, 0), 2),
            })
    df_routes = pd.DataFrame(route_rows) if route_rows else pd.DataFrame()

    # ── Write Excel ──────────────────────────────────────────
    xlsx_path = out / f"{instance_name}_pareto.xlsx"
    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Frontier", index=False)
        if not df_routes.empty:
            df_routes.to_excel(writer, sheet_name="Routes", index=False)

    # Highlight knee point in amber
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb = load_workbook(str(xlsx_path))
        ws = wb["Frontier"]
        amber = PatternFill(start_color="FFBF00", end_color="FFBF00",
                            fill_type="solid")
        for row_idx, sol in enumerate(frontier):
            if sol.get("knee", False):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx + 2, column=col).fill = amber
        wb.save(str(xlsx_path))
    except Exception:
        pass

    print(f"  ✓ {xlsx_path}")

    # ── Write JSON ───────────────────────────────────────────
    json_path = out / f"{instance_name}_pareto.json"
    json_data = []
    for idx, sol in enumerate(frontier):
        entry = {
            "point": idx + 1,
            "W1": sol["W1"],
            "W2": sol["W2"],
            "status": sol["status"],
            "knee": sol.get("knee", False),
            "L": {str(k): v for k, v in sol.get("L", {}).items()},
            "routes": {str(k): v for k, v in sol.get("routes", {}).items()},
        }
        json_data.append(entry)

    with open(json_path, "w") as fp:
        json.dump(json_data, fp, indent=2, default=str)
    print(f"  ✓ {json_path}")
